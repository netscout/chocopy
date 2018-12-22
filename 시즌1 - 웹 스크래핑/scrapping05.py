import requests as req
import bs4
import openpyxl as xl


def getCourseList(url):
    #웹 페이지의 내용을 읽어와야
    res = req.get(url)

    #읽어온 웹 페이지의 내용을 출력
    #print(res.text)

    #특정 요소(강좌)만 검색하기
    bs = bs4.BeautifulSoup(res.text, features="html.parser")

    courses = bs.select(".track__course")
    
    #강좌를 저장할 강좌 목록
    courseList = []

    for c in courses:
        link = ""
        linkTag = c.select_one("a")
        if linkTag is None:
            break
        link = linkTag.attrs["href"]
        title = c.select_one("h4.course-block__title").getText().strip()
        desc = c.select_one("p.course-block__description").getText().strip()

        courseList.append({"link":link, "title":title, "desc":desc})
    return courseList

def getChapterList(url):
    courseUrl = "https://www.datacamp.com" + url

    courseRes = req.get(courseUrl)

    courseBs = bs4.BeautifulSoup(courseRes.text, features="html.parser")

    #강좌 페이지에서 챕터 목록을 검색
    chapters = courseBs.select("ol.chapters > li.chapter")
    chapterList = []

    for chap in chapters:
        title = chap.select_one("h4.chapter__title").getText().strip()
        desc = chap.select_one("p.chapter__description").getText().strip()
        print("...")

        subChapterTitles = chap.select(".chapter__exercise-title")
        subChapterTitleList = []
        
        for sc in subChapterTitles:
            subTitle = sc.getText().strip()
            subChapterTitleList.append(subTitle)

        chapterList.append({"title":title, "desc":desc, "subChapterTitles":subChapterTitleList})

    return chapterList

def saveToExcel(filename, courseList):
    wb = xl.Workbook()
    sheet = wb.active
    sheet.title = "데이터 사이언티스트 강좌"

    row = 2
    col = 1
    for c in courseList:
        sheet.cell(row = row, column = col).value = c["link"]
        sheet.cell(row = row, column = col + 1).value = c["title"]
        sheet.cell(row = row, column = col + 2).value = c["desc"]
        row = row + 1
        for chap in c["chapters"]:
            sheet.cell(row = row, column = col + 1).value = chap["title"]
            sheet.cell(row = row, column = col + 2).value = chap["desc"]
            row = row + 1
            for scTitle in chap["subChapterTitles"]:
                sheet.cell(row = row, column = col + 2).value = scTitle
                row = row + 1

    wb.save(filename)

#읽어올 웹 페이지의 주소가 필요
url = "https://www.datacamp.com/tracks/data-scientist-with-python"

#강좌를 저장할 강좌 목록
courseList = []

#강좌 목록을 읽어오기
courseList = getCourseList(url)

#세부 챕터 읽어오기
for c in courseList:
    chapterList = getChapterList(c["link"])
    c["chapters"] = chapterList

#엑셀로 읽어온 내용 저장하기
saveToExcel("DS with Python.xlsx", courseList)
