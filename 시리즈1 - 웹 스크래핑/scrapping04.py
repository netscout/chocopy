import requests as req
import bs4
import openpyxl as xl

#읽어올 페이지 주소 필요
url = "https://www.datacamp.com/tracks/data-scientist-with-python"

#주소에 해당하는 웹 페이지 읽어오기
res = req.get(url)

#특정 요소만 검색하기
bs = bs4.BeautifulSoup(res.text, features="html.parser")

courses = bs.select(".track__course")
courseList = []

#검색한 요소 출력하기
for c in courses:
    link = ""
    linkTag = c.select_one("a")
    if linkTag is None:
        break;
    link = linkTag.attrs["href"]
    title = c.select_one("h4.course-block__title").getText().strip()
    desc = c.select_one("p.course-block__description").getText().strip()
    courseList.append({"link":link, "title":title, "desc":desc})
    
##for c in courseList:
##    print(c["link"], c["title"], c["desc"])

#검색한 요소(강좌)를 목록으로 저장하기



#각 강좌의 웹 페이지를 가져오기
for c in courseList:
    #각 강좌의 챕터 목록 검색하기(제목, 설명)
    courseUrl = c["link"]

    courseRes = req.get("https://www.datacamp.com" + courseUrl)

    courseBs = bs4.BeautifulSoup(courseRes.text, features="html.parser")

    chapters = courseBs.select("ol.chapters > li.chapter")

    chapterList = []

    for chap in chapters:
        title = chap.select_one(".chapter__title").getText().strip()
        desc = chap.select_one(".chapter__description").getText().strip()
        #print("title : ", title)

        #각 챕터의 세무 챕터목록 검색하기(제목)
        subChapterTitles = chap.select(".chapter__exercise-title")
        subChapterTitleList = []
        
        for sc in subChapterTitles:
            subTitle = sc.getText().strip()
            subChapterTitleList.append(subTitle)
            #print("subTitle : ", subTitle)

        chapterList.append({"title":title, "desc":desc, "subChapterTitles":subChapterTitleList})
        print("...")

    c["chapters"] = chapterList

wb = xl.Workbook()
#기본 시트 가져오기
sheet = wb.active
#시트 제목 설정
sheet.title = "Data Scientist with Python"

rowNum = 2
colNum = 1
for c in courseList:
    sheet.cell(row=rowNum, column=colNum).value = c["link"]
    sheet.cell(row=rowNum, column=colNum + 1).value = c["title"]
    sheet.cell(row=rowNum, column=colNum + 2).value = c["desc"]
    rowNum = rowNum + 1
    for chap in c["chapters"]:
        sheet.cell(row=rowNum, column=colNum + 1).value = chap["title"]
        sheet.cell(row=rowNum, column=colNum + 2).value = chap["desc"]
        rowNum = rowNum + 1
        for scTitle in chap["subChapterTitles"]:
            sheet.cell(row=rowNum, column=colNum + 2).value = scTitle
            rowNum = rowNum + 1

wb.save("Data Scientist with Python.xlsx")            

