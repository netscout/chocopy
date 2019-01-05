import openpyxl as xl
import os

currentPath = os.path.dirname(__file__)
salesPath = os.path.join(currentPath, "매출현황")

filePath = os.path.join(salesPath, "거래처 매출 현황.xlsx")

sales = {}

wb = xl.load_workbook(filePath)

for sheetname in wb.sheetnames:
    sheet = wb[sheetname]
    # 0  1  2  3  4
    #[0, 0, 0, 0, 0, ...12개]
    values = [0] * 12

    for month in range(1, 13):
        cellValue = sheet.cell(row=month+1, column=2).value
        if cellValue is None:
            cellValue = 0

        #values 리스트의 값 중에 month-1 번째 공간에 cellValue의 값을 저장
        #month = 1, month-1 = 0
        #리스트의 1번째 값을 가져오기 위해서는 0번째 값을 접근해야
        values[month-1] = cellValue

    #key:"거래처1" - [11000, 12000, .....]
    sales[sheetname] = values

sheet = wb.create_sheet("합계")

#헤더 부분
col = 1
headers = ["거래처명", "1월", "2월", "3월", "4월", "5월", "6월", "7월", "8월", "9월", "10월", "11월", "12월", "합계"]
for header in headers:
    sheet.cell(row=1, column=col).value = header
    col += 1

#각 거래처별 매출 정보
row = 2
for key, values in sales.items():
    sheet.cell(row=row, column=1).value = key

    col = 2
    for value in values:
        sheet.cell(row=row, column=col).value = value
        col += 1  
        
    #fromCell = xl.utils.get_column_letter(2) + str(row)
    #B2
    fromCell = f"{xl.utils.get_column_letter(2)}{row}"
    toCell = f"{xl.utils.get_column_letter(col-1)}{row}"

    sheet.cell(row=row, column=col).value = f"=SUM({fromCell}:{toCell})"
    
    row += 1
    #=SUM(B2:B13)

#맨 아래쪽에 합계 행

sheet.cell(row = row, column = 1).value = "합계"
#2열 부터 14열까지
for i in range(2, 15):
    fromCell = f"{xl.utils.get_column_letter(i)}2"
    toCell = f"{xl.utils.get_column_letter(i)}{row-1}"
    sheet.cell(row=row, column=i).value = f"=SUM({fromCell}:{toCell})"

#헤더 행과 열에 대해서 스타일을 입히기
bold = xl.styles.Font(bold=True)
center = xl.styles.Alignment(horizontal="center")
bd = xl.styles.Side(style="thin", color="000000")
border = xl.styles.Border(left=bd, top=bd, right=bd, bottom=bd)

#헤더 스타일
headerStyle = xl.styles.NamedStyle(name="header")
headerStyle.font = bold
headerStyle.border = border
headerStyle.alignment = center
#header
wb.add_named_style(headerStyle)

#일반 셀 스타일
normalStyle = xl.styles.NamedStyle(name="normal")
normalStyle.border = border
#normal
wb.add_named_style(normalStyle)

for row in sheet.rows:
    for cell in row:
        if cell.row == 1 or cell.column == "A":
            cell.style = "header"
        else:
            cell.style = "normal"

wb.save(filePath)
