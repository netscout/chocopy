import openpyxl as xl
import os

#c:\sources
currentPath = os.path.dirname(__file__)
#c:\sources + \리포트 =>
reportPath = os.path.join(currentPath, "리포트")

reports = []

for file in os.listdir(reportPath):
    if file.endswith(".xlsx") and "20181210" in file:
        filePath = os.path.join(reportPath, file)
        
        wb = xl.load_workbook(filePath)
        sheet = wb.active

        #3행부터 1열-5열
        #name, lastweek, thisweek, nextweek, issue
        row = 3
        col = 1

        #10번 반복하겠다(팀원이 10명을 넘지 않는다는 가정)
        for i in range(10):
            name = sheet.cell(row = row, column = col).value
            if name is None:
                break
            lastweek = sheet.cell(row = row, column = col+1).value
            thisweek = sheet.cell(row = row, column = col+2).value
            nextweek = sheet.cell(row = row, column = col+3).value
            issue = sheet.cell(row = row, column = col+4).value

            reports.append({"name":name, "lastweek":lastweek, "thisweek":thisweek, "nextweek":nextweek, "issue":issue})
            row += 1

#지난주 리포트를 읽어오고
wb = xl.load_workbook(os.path.join(reportPath, "Weekly Report_20181203.xlsx"))
sheet = wb.active

nameList = {} #[] {}
newRow = 0

for i in range(3, 13):
    name = sheet.cell(row = i, column = 1).value
    if name is None:
        newRow = i
        break
    #key : value
    #"홍길동" : 3
    #"영희" : 4
    nameList[name] = i

#지난 주 리포트에 저장되어 있는 팀원들의 위치 정보를 확인

#읽어온 데이터를 바탕으로 새로 데이터를 작성한다
col = 1
for r in reports:
    row = nameList.get(r["name"], -1)
    if row == -1:
        row = newRow
        newRow += 1
    
    sheet.cell(row = row, column = col).value = r["name"]
    sheet.cell(row = row, column = col+1).value = r["lastweek"]
    sheet.cell(row = row, column = col+2).value = r["thisweek"]
    sheet.cell(row = row, column = col+3).value = r["nextweek"]
    sheet.cell(row = row, column = col+4).value = r["issue"]

wb.save(os.path.join(reportPath, "Weekly Report_20181210.xlsx"))